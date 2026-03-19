from datetime import datetime, timedelta, timezone
import io
import pandas as pd
import streamlit as st

# Singapore Standard Time (UTC+8) — used for all display timestamps
SGT = timezone(timedelta(hours=8))

from config import CATEGORIES
from news_fetcher import fetch_all_categories, fetch_articles_for_category, _get_window_hours
from llm_processor import enrich_articles

# ---------------------------------------------------------------------------
# Page config
# ---------------------------------------------------------------------------
st.set_page_config(
    page_title="Helicap News",
    page_icon="🔵",
    layout="wide",
)

# ---------------------------------------------------------------------------
# Custom CSS — navy/white/blue palette + Google Fonts
# ---------------------------------------------------------------------------
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;600;700&family=Inter:wght@400;500&display=swap');

html, body, [class*="css"] {
    font-family: 'Inter', sans-serif;
    background-color: #FFFFFF;
    color: #111827;
}

/* ── Header ─────────────────────────────────────────────────────────────── */
.helicap-header {
    background-color: #0D1F3C;
    padding: 1.2rem 2rem;
    border-radius: 8px;
    display: flex;
    align-items: center;
    justify-content: space-between;
    margin-bottom: 1.5rem;
}
.helicap-header h1 {
    font-family: 'DM Sans', sans-serif;
    color: #FFFFFF;
    font-size: 1.8rem;
    font-weight: 700;
    margin: 0 0 0.15rem 0;
    display: flex;
    align-items: center;
    gap: 0.45rem;
}
/* Tagline indented to align with the "H" in Helicap (skips icon + gap) */
.header-tagline {
    color: #7B93B8;
    font-size: 0.95rem;
    font-style: italic;
    font-family: 'Inter', sans-serif;
    padding-left: 2.35rem;   /* ≈ emoji width (1.8rem) + gap (0.45rem) + tiny buffer */
}

/* ── Sidebar credit block ────────────────────────────────────────────────── */
.sidebar-credit {
    background-color: #0D1F3C;
    border-radius: 8px;
    padding: 1.2rem 1rem 1rem;
    margin-bottom: 0.6rem;
    text-align: center;
}
.sidebar-credit .credit-label {
    color: #7B93B8;
    font-size: 0.68rem;
    text-transform: uppercase;
    letter-spacing: 0.09em;
    margin-bottom: 0.35rem;
}
.sidebar-credit .credit-name {
    color: #FFFFFF;
    font-family: 'DM Sans', sans-serif;
    font-size: 1.0rem;       /* ↓ slightly smaller than before */
    font-weight: 700;
    line-height: 1.2;
}

/* ── Sidebar last-updated ────────────────────────────────────────────────── */
.sidebar-updated {
    color: #6B7280;
    font-size: 0.88rem;      /* ↑ more readable */
    text-align: center;
    line-height: 1.7;
    margin-bottom: 0.25rem;
}
.sidebar-updated strong {
    font-size: 0.92rem;
    color: #374151;
}

/* ── Category badge ──────────────────────────────────────────────────────── */
.category-badge {
    display: inline-block;
    background-color: #0D1F3C;
    color: #FFFFFF;
    font-family: 'DM Sans', sans-serif;
    font-weight: 600;
    font-size: 0.9rem;
    padding: 0.35rem 0.9rem;
    border-radius: 20px;
    margin-bottom: 0.75rem;
}

/* ── Article card ────────────────────────────────────────────────────────── */
.article-card {
    background-color: #EBF2FF;
    border-radius: 8px;
    padding: 1rem 1.25rem;
    margin-bottom: 0.75rem;
    border-left: 4px solid #1A56DB;
}
.article-title {
    font-family: 'DM Sans', sans-serif;
    font-weight: 600;
    font-size: 1rem;
    color: #111827;
    margin-bottom: 0.25rem;
}
.article-meta {
    font-size: 0.78rem;
    color: #6B7280;
    margin-bottom: 0.5rem;
}
.article-summary {
    font-size: 0.88rem;
    color: #374151;
    margin-bottom: 0.6rem;
    line-height: 1.5;
}
.sentiment-positive { color: #F0A500; font-weight: 600; font-size: 0.82rem; }
.sentiment-negative { color: #E63946; font-weight: 600; font-size: 0.82rem; }
.sentiment-neutral  { color: #6B7280; font-weight: 600; font-size: 0.82rem; }

.no-articles {
    color: #6B7280;
    font-size: 0.88rem;
    font-style: italic;
    padding: 0.5rem 0;
}

/* ── Read link ───────────────────────────────────────────────────────────── */
.read-link a {
    color: #1A56DB;
    font-size: 0.85rem;
    font-weight: 500;
    text-decoration: none;
}
.read-link a:hover { text-decoration: underline; }

/* ── Buttons ─────────────────────────────────────────────────────────────── */
div[data-testid="stButton"] > button {
    background-color: #1A56DB;
    color: #FFFFFF;
    border: none;
    border-radius: 6px;
    font-family: 'DM Sans', sans-serif;
    font-weight: 600;
    padding: 0.5rem 1.5rem;
    width: 100%;
}
div[data-testid="stButton"] > button:hover {
    background-color: #1446B8;
}

/* ── Footer ──────────────────────────────────────────────────────────────── */
.helicap-footer {
    text-align: center;
    color: #6B7280;
    font-size: 0.78rem;
    padding: 1.5rem 0 0.5rem;
    border-top: 1px solid #E5E7EB;
    margin-top: 2rem;
}
</style>
""", unsafe_allow_html=True)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _format_timestamp(iso_str: str) -> str:
    try:
        dt = datetime.fromisoformat(iso_str.replace("Z", "+00:00"))
        return dt.astimezone(SGT).strftime("%b %d, %Y · %I:%M %p SGT")
    except Exception:
        return iso_str


def _sentiment_html(sentiment: str) -> str:
    s = sentiment.strip().capitalize()
    css = {
        "Positive": "sentiment-positive",
        "Negative": "sentiment-negative",
        "Neutral":  "sentiment-neutral",
    }.get(s, "sentiment-neutral")
    icon = {"Positive": "🟡", "Negative": "🔴", "Neutral": "⚪"}.get(s, "⚪")
    return f'<span class="{css}">{icon} {s}</span>'


def _render_article_card(article: dict) -> None:
    title     = article.get("title", "")
    source    = article.get("source", "")
    pub       = _format_timestamp(article.get("published_at", ""))
    summary   = article.get("summary", "") or "Summary unavailable."
    sentiment = article.get("sentiment", "Neutral")
    url       = article.get("url", "#")
    trusted_badge = " ✓" if article.get("trusted") else ""

    st.markdown(f"""
    <div class="article-card">
        <div class="article-title">📰 {title}</div>
        <div class="article-meta">{source}{trusted_badge} · {pub}</div>
        <div class="article-summary">{summary}</div>
        <div style="display:flex; justify-content:space-between; align-items:center;">
            {_sentiment_html(sentiment)}
            <span class="read-link"><a href="{url}" target="_blank">Read Article →</a></span>
        </div>
    </div>
    """, unsafe_allow_html=True)


def _build_dataframe(results: dict) -> pd.DataFrame:
    rows = []
    for category, articles in results.items():
        for a in articles:
            rows.append({
                "category":     category,
                "title":        a.get("title", ""),
                "source":       a.get("source", ""),
                "published_at": a.get("published_at", ""),
                "url":          a.get("url", ""),
                "summary":      a.get("summary", ""),
                "sentiment":    a.get("sentiment", ""),
            })
    return pd.DataFrame(rows)


def _build_newsletter_excel(results: dict, run_date: str) -> bytes:
    """
    Generate a formatted Excel workbook matching the newsletter template:
    - Category header rows (navy background, white bold text)
    - One row per article: bullet title in col A, clickable URL in col B
    Returns the file as bytes for st.download_button.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    NAVY   = "0D1F3C"
    BLUE   = "D6E4F7"   # light blue for alternating article rows
    WHITE  = "FFFFFF"

    wb = Workbook()
    ws = wb.active
    ws.title = "Helicap News"

    # Column widths
    ws.column_dimensions["A"].width = 80
    ws.column_dimensions["B"].width = 45

    # Top header
    ws.append([f"Helicap News  ·  {run_date}", ""])
    hdr = ws[1]
    for cell in hdr:
        cell.font      = Font(name="Calibri", bold=True, size=13, color=WHITE)
        cell.fill      = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 22
    ws.merge_cells("A1:B1")
    ws.append([])   # blank spacer

    row_idx = 3
    for category, articles in results.items():
        real = [a for a in articles if "_error" not in a and a.get("title")]
        # Category header
        cat_label = category.replace("(", "").replace(")", "").strip()
        ws.cell(row=row_idx, column=1, value=cat_label)
        ws.cell(row=row_idx, column=2, value="")
        for col in (1, 2):
            c = ws.cell(row=row_idx, column=col)
            c.font      = Font(name="Calibri", bold=True, size=11, color=WHITE)
            c.fill      = PatternFill("solid", fgColor="1A3A5C")
            c.alignment = Alignment(vertical="center")
        ws.row_dimensions[row_idx].height = 18
        row_idx += 1

        if not real:
            ws.cell(row=row_idx, column=1, value="   No recent articles found in the last 24 hours.")
            ws.cell(row=row_idx, column=1).font = Font(name="Calibri", italic=True, color="6B7280")
            row_idx += 1
        else:
            fill_toggle = False
            for a in real:
                title = a.get("title", "")
                url   = a.get("url", "#")
                fill_bg = BLUE if fill_toggle else WHITE

                title_cell = ws.cell(row=row_idx, column=1, value=f"  •  {title}")
                title_cell.font      = Font(name="Calibri", size=10)
                title_cell.fill      = PatternFill("solid", fgColor=fill_bg)
                title_cell.alignment = Alignment(wrap_text=True, vertical="center")

                link_cell = ws.cell(row=row_idx, column=2, value="Open article →")
                link_cell.hyperlink  = url
                link_cell.font       = Font(name="Calibri", size=10, color="1A56DB", underline="single")
                link_cell.fill       = PatternFill("solid", fgColor=fill_bg)
                link_cell.alignment  = Alignment(vertical="center")

                ws.row_dimensions[row_idx].height = 28
                fill_toggle = not fill_toggle
                row_idx += 1

        row_idx += 1   # blank row between categories

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Sidebar
# ---------------------------------------------------------------------------
all_category_keys = list(CATEGORIES.keys())

with st.sidebar:
    # Credit block — prominent, navy-styled
    st.markdown("""
    <div class="sidebar-credit">
        <div class="credit-label">Created and Designed by</div>
        <div class="credit-name">Laura Garzon</div>
    </div>
    """, unsafe_allow_html=True)

    # Last updated timestamp
    last_updated = st.session_state.get("last_updated", "—")
    st.markdown(
        f'<div class="sidebar-updated">Last updated<br><strong>{last_updated}</strong></div>',
        unsafe_allow_html=True,
    )

    st.divider()

    # Category filter — resets to all on fresh page load
    selected_categories = st.multiselect(
        "Categories",
        options=all_category_keys,
        default=all_category_keys,
        help="Deselect categories you don't need — saves API quota.",
    )

    # Articles per category
    max_articles = st.slider(
        "Articles per category",
        min_value=1,
        max_value=5,
        value=2,
        help="Niche categories may have fewer articles available in the last 24 h.",
    )

    st.divider()

    refresh_clicked = st.button("🔄 Refresh News")

    # Debug mode — collapsed by default
    st.divider()
    with st.expander("🛠 Debug Mode"):
        st.caption("Test a single category without a full refresh or AI call.")
        debug_category = st.selectbox(
            "Category to test",
            options=all_category_keys,
            key="debug_category_select",
        )
        if st.button("🔍 Test this category", key="debug_fetch_btn"):
            window_hours = _get_window_hours()
            window_start = datetime.now(timezone.utc) - timedelta(hours=window_hours)
            with st.spinner(f"Fetching '{debug_category}'…"):
                try:
                    debug_articles = fetch_articles_for_category(
                        debug_category, window_start, max_articles
                    )
                except Exception as exc:
                    st.error(f"Error: {exc}")
                    debug_articles = []
            if not debug_articles:
                st.warning(f"No articles found in the last {window_hours} h.")
            else:
                st.success(f"{len(debug_articles)} article(s) — raw data, no AI summary.")
                for a in debug_articles:
                    trusted = "✓ trusted" if a.get("trusted") else "· unverified"
                    pub = a.get("published_at", "")[:16].replace("T", " ")
                    st.markdown(f"""
**{a['title']}**
`{a['source']}` {trusted} · `{pub}`
{a.get('url', '')}
---""")


# ---------------------------------------------------------------------------
# Header
# ---------------------------------------------------------------------------
st.markdown("""
<div class="helicap-header">
    <div>
        <h1><span>🔵</span><span>Helicap News</span></h1>
        <div class="header-tagline">Stay informed. Stay ahead.</div>
    </div>

</div>
""", unsafe_allow_html=True)


# ---------------------------------------------------------------------------
# Refresh logic
# ---------------------------------------------------------------------------
if refresh_clicked:
    if not selected_categories:
        st.warning("Please select at least one category.")
        st.stop()
    with st.spinner("Fetching latest articles and generating summaries…"):
        try:
            raw_results, window_hours = fetch_all_categories(selected_categories, max_articles)
        except RuntimeError as e:
            err_msg = str(e).lower()
            if any(k in err_msg for k in ("quota exhausted", "429", "426", "ratelimited")):
                st.error(
                    "⚠️ NewsAPI daily quota exhausted (100 req/day on free tier). "
                    "GNews fallback also unavailable. Please try again tomorrow or "
                    "add a GNEWS_API_KEY to your secrets."
                )
            else:
                st.error(f"News fetch failed: {e}")
            st.stop()

        all_articles = [
            a
            for articles in raw_results.values()
            for a in articles
            if "_error" not in a
        ]

        if all_articles:
            try:
                enrich_articles(all_articles)
            except RuntimeError as e:
                st.warning(f"AI summaries unavailable: {e}. Showing articles without summaries.")

        st.session_state["results"]      = raw_results
        st.session_state["window_hours"] = window_hours
        st.session_state["last_updated"] = datetime.now(SGT).strftime("%B %d, %Y · %H:%M SGT")
        st.rerun()


# ---------------------------------------------------------------------------
# Results
# ---------------------------------------------------------------------------
results: dict = st.session_state.get("results", {})

if not results:
    st.info("Click **Refresh News** in the sidebar to load the latest articles.")
else:
    window_hours = st.session_state.get("window_hours", 24)
    if window_hours > 24:
        st.info(
            f"🌙 Weekend / off-hours mode · Showing analysis, previews & forecasts "
            f"· Last {window_hours} hours"
        )

    for category in all_category_keys:
        if category not in results:
            continue
        articles = results[category]

        st.markdown(f'<div class="category-badge">{category}</div>', unsafe_allow_html=True)

        if not articles:
            wh = st.session_state.get("window_hours", 24)
            st.markdown(
                f'<div class="no-articles">No recent articles found in the last {wh} hours.</div>',
                unsafe_allow_html=True,
            )
        elif len(articles) == 1 and "_error" in articles[0]:
            err_lower = articles[0]["_error"].lower()
            if any(k in err_lower for k in ("quota exhausted", "429", "426", "ratelimited")):
                st.markdown(
                    '<div class="no-articles">⚠️ API quota exhausted for this category. Try again later.</div>',
                    unsafe_allow_html=True,
                )
            else:
                st.markdown(
                    f'<div class="no-articles">⚠️ Could not fetch: {articles[0]["_error"]}</div>',
                    unsafe_allow_html=True,
                )
        else:
            for article in articles:
                _render_article_card(article)

        st.markdown("<br>", unsafe_allow_html=True)

    # Downloads
    run_date = datetime.now(SGT).strftime("%Y-%m-%d")
    col_xl, col_csv = st.columns(2)
    with col_xl:
        st.download_button(
            label="⬇ Download Newsletter (Excel)",
            data=_build_newsletter_excel(results, run_date),
            file_name=f"helicap_news_{run_date}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    with col_csv:
        df = _build_dataframe(results)
        st.download_button(
            label="⬇ Download Full CSV",
            data=df.to_csv(index=False).encode("utf-8"),
            file_name=f"helicap_news_{run_date}.csv",
            mime="text/csv",
        )


# ---------------------------------------------------------------------------
# Footer
# ---------------------------------------------------------------------------
st.markdown("""
<div class="helicap-footer">
    Curated by AI · Designed by Laura Garzon · 2026
</div>
""", unsafe_allow_html=True)
